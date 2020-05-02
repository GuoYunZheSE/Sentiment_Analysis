# @Date    : 16:52 05/01/2020
# @Author  : ClassicalPi
# @FileName: Data_Preprocessing.py
# @Software: PyCharm

import numpy
import pandas
import openpyxl
import os

def getStaticInformation(City:str):
    os.chdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(City))
    files = os.listdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(City))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "{}".format(City)
    ws.append(["餐厅名称", "评论数量"])
    for file in files:
        if file.endswith(".xlsx"):
            print(file)
            wb2 = openpyxl.load_workbook(file)
            ws2 = wb2.active
            file = file.split('.')[0]
            name = file.replace("_", " ")
            number = ws2.max_row - 1
            ws.append([name, number])
            wb2.close()
    wb.save("{}.xlsx".format("{}_Static_Information".format(City)))
    print("Done")

if __name__ == '__main__':
    getStaticInformation("Macau")