# @Date    : 22:38 05/01/2020
# @Author  : ClassicalPi
# @FileName: Data_Analysis.py
# @Software: PyCharm

import numpy as np
import pandas as pd
import nltk
import re
import os
import matplotlib.pyplot as plt
import codecs
import openpyxl
import mpld3
import matplotlib
from nltk.stem.snowball import SnowballStemmer
from sklearn import feature_extraction

def loadReview(city:str, res:None, All:False)->str:
    res=""
    os.chdir("/Users/lucas/Projects/Pycharm/Sentiment_Analysis/Data/{}".format(city))
    if All:
        wb = openpyxl.load_workbook("All_{}.xlsx".format(city))
    else:
        wb = openpyxl.load_workbook("{}.xlsx".format(res))
    ws = wb.active
    for row in range(1,ws.max_row):
        temp=str(ws.cell(row=row,column=5).value)
        if ord(temp[0])>=65 and ord(temp[0])<=122:
            res+=temp
            res+='\n'
    return res

def tokenize_and_stem(text):
    # first tokenize by sentence, then by word to ensure that punctuation is caught as it's own token
    tokens = [word for sent in nltk.sent_tokenize(text) for word in nltk.word_tokenize(sent)]
    filtered_tokens = []
    # filter out any tokens not containing letters (e.g., numeric tokens, raw punctuation)
    for token in tokens:
        if re.search('[a-zA-Z]', token):
            filtered_tokens.append(token)
    stems = [stemmer.stem(t) for t in filtered_tokens]
    return stems


def tokenize_only(text):
    # first tokenize by sentence, then by word to ensure that punctuation is caught as it's own token
    tokens = [word.lower() for sent in nltk.sent_tokenize(text) for word in nltk.word_tokenize(sent)]
    filtered_tokens = []
    # filter out any tokens not containing letters (e.g., numeric tokens, raw punctuation)
    for token in tokens:
        if re.search('[a-zA-Z]', token):
            filtered_tokens.append(token)
    return filtered_tokens

def drawBarh(dic:dict,num:int):
    """
    绘制水平条形图方法barh
    参数一：y轴
    参数二：x轴
    """
    listkey=[]
    listval=[]
    for key, val in sorted(dic.items(), key=lambda x: (x[1], x[0]), reverse=True)[:num]:
        listkey.append(key)
        listval.append(val)
    df = pd.DataFrame(listval, columns=[u'Times'])
    df.index = listkey
    df.plot(kind='barh',color="purple")
    plt.title(u'Top {} Most Common Words in Reviews'.format(num))
    plt.show()


if __name__ == '__main__':
    synopses=[loadReview("GuangZhou",All=True,res=None)]
    stopwords = nltk.corpus.stopwords.words('english')
    stopwords.append('\'s')
    stopwords.append('n\'t')
    stopwords.append('us')
    stemmer = SnowballStemmer("english")
    totalvocab_stemmed = []
    totalvocab_tokenized = []
    for i in synopses:
        allwords_stemmed = tokenize_and_stem(i)  # for each item in 'synopses', tokenize/stem
        totalvocab_stemmed.extend(allwords_stemmed)  # extend the 'totalvocab_stemmed' list

        allwords_tokenized = tokenize_only(i)
        totalvocab_tokenized.extend(allwords_tokenized)
    words=[]
    for each in totalvocab_tokenized:
        if each not in stopwords:
            words.append(each)
    fre=nltk.FreqDist(words)
    drawBarh(fre,20)